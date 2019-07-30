import numpy as np
import matplotlib.pyplot as plt
import openpyxl as op
import math 

def parse_data(file_name,target,bound,data_type=int,*kwargs):
    '''A utility to remove a percentage of data from an excel workbook
    can be run directly from your IDE of choice(developed in Spyder,Python3.7)
    Takes as arguements:
        Arguement:  Type:      Description:
        file_name:   String       file name NOT INCLUDING extension for excel, 
                                  can be relavtive or absolute
                                  
        target:      float/int    target value to be used for removal of data
        
        bound:       float        percent of data to be removed IN DECIMAL FORM
        
        data_type:   object       either int or float, depending on data 
                                  included in the spreadsheet for 
                                  parsing data correctly in final stages
    '''
    #convert the excel document into something easier to work with
    book=op.load_workbook(file_name+'.xlsx',guess_types=True)
    sheet=book.active
    data=sheet.iter_rows()
    #put the data from the spread sheet into a 2D matrix
    matrix=[[row[i].value for i in range(len(row))] for row in data]
    
    #convert the 2D matrix into a 1D matrix of absolute values to determine
    #the bound region, will remove all the NoneTypes for only this matrix
    #also get the number of actual entries to compare to for the bounds
    matrix_1D=[abs(target-d) for u in matrix for d in u if type(d)==data_type]
    number_of_values=len(matrix_1D)
    
    #sort the 1D matrix into ascending order and get rid of the final 20%
    #this might remove 1 value less than desired. IF this is an issue
    #change math.trunc() to int()
    matrix_1D_sorted=np.sort(matrix_1D)[:math.trunc(-number_of_values*bound)]
    #get the maxiumum tolerable value
    max_abs_value=matrix_1D_sorted[-1]
    
    #take the absolute value of each element and compare it the max tolerance
    #form a 2D matrix with absolute values for use in plotting it later
    matrix_2D=matrix.copy()
    for j in range(len(matrix)):
        for k in range(len(matrix[j])):
            if type(matrix[j][k])==data_type:
                if abs(matrix[j][k]-target)>max_abs_value:
                    matrix[j][k]=None
                    matrix_2D[j][k]=np.nan
                else:
                    matrix_2D[j][k]=abs(matrix[j][k]-target)
            else:
                matrix_2D[j][k]=np.nan
                    
#    create a new sheet in the excel workbook
    base_name='Parsed_Data_{}_{}'.format(target,bound)
    book.create_sheet(base_name)
    new_sheet=book[base_name]
    #write the data to the spread sheet skipping all the instances 
    #that have been removed
    for h in range(len(matrix)):
        for l in range(len(matrix[h])):
            if type(matrix[h][l])==data_type:
                new_sheet.cell(row=h+1,column=l+1).value=matrix[h][l]
    #saving the file
    book.save(file_name+'.xlsx')
    book.close()
    return matrix_2D

a=parse_data('RAW_DATA',250,0.2)

fig = plt.figure(figsize=(10, 10))
plt.imshow(a,interpolation='None')
plt.title('Variance from target')
plt.xlabel('Row')
plt.ylabel('Column')
plt.colorbar()
plt.show()